import openpyxl
from openpyxl.styles import Alignment

if __name__ == '__main__':
    workbook = openpyxl.load_workbook("bilibili_bangumi.xlsx")
    # master_sheet = workbook["0"]
    # master_row_count = master_sheet.max_row
    # master_row_count += 1
    # # for i in range(2, len(workbook.sheetnames)):
    # # slave_sheet = workbook.get_sheet_by_name(workbook.sheetnames[i])
    # slave_sheet = workbook["15000"]
    # slave_row_count = slave_sheet.max_row
    # for j in range(1, slave_row_count):
    #     master_sheet.cell(master_row_count, 1).value = slave_sheet.cell(j + 1, 1).value
    #     master_sheet.cell(master_row_count, 1).alignment = Alignment(horizontal='center', vertical='center')
    #     master_sheet.cell(master_row_count, 2).value = slave_sheet.cell(j + 1, 2).value
    #     master_sheet.cell(master_row_count, 2).alignment = Alignment(horizontal='center', vertical='center')
    #     master_sheet.cell(master_row_count, 3).value = slave_sheet.cell(j + 1, 3).value
    #     master_sheet.cell(master_row_count, 3).alignment = Alignment(horizontal='center', vertical='center')
    #     master_row_count += 1
    # workbook.remove(slave_sheet)
    sheet = workbook.create_sheet(title="100000")
    sheet.column_dimensions['B'].width = 50
    sheet.cell(1, 1).value = "id"
    sheet.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(1, 2).value = "name"
    sheet.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')
    sheet.cell(1, 3).value = "episode"
    sheet.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
    workbook.save("bilibili_bangumi.xlsx")
