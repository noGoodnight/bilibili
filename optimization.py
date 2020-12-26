import time

import openpyxl
import requests
from openpyxl.styles import Alignment


def get_content(ep_id: int) -> str:
    data = requests.get("https://www.bilibili.com/bangumi/play/ep" + str(ep_id))
    content = data.content.decode("utf8")
    if "由于触发哔哩哔哩安全风控策略，该次访问请求被拒绝。" in content:
        return "412"
    else:
        return content


def save_series(sheet, series_info):
    if sheet.cell(series_info[0], 1).value is not None:
        return
    sheet.cell(series_info[0], 1).value = series_info[0]
    sheet.cell(series_info[0], 2).value = series_info[1]
    sheet.cell(series_info[0], 3).value = series_info[2]
    sheet.cell(series_info[0], 4).value = series_info[3]
    sheet.cell(series_info[0], 5).value = series_info[4]
    sheet.cell(series_info[0], 6).value = series_info[5]
    sheet.cell(series_info[0], 7).value = series_info[6]
    sheet.cell(series_info[0], 8).value = series_info[7]
    sheet.cell(series_info[0], 9).value = series_info[8]
    sheet.cell(series_info[0], 10).value = series_info[9]
    sheet.cell(series_info[0], 11).value = series_info[10]
    sheet.cell(series_info[0], 12).value = series_info[11]
    sheet.cell(series_info[0], 13).value = series_info[12]
    sheet.cell(series_info[0], 14).value = series_info[13]
    sheet.cell(series_info[0], 15).value = series_info[14]
    sheet.cell(series_info[0], 16).value = series_info[15]

    print(series_info[0], end="  ")
    print(series_info[1], end="  ")
    print(series_info[2], end="  ")
    print(series_info[3], end="  ")
    print(series_info[4], end="  ")
    print(series_info[5], end="  ")
    print(series_info[6], end="  ")
    print(series_info[7], end="  ")
    print(series_info[8], end="  ")
    print(series_info[9], end="  ")
    print(series_info[10], end="  ")
    print(series_info[11], end="  ")
    print(series_info[12], end="  ")
    print(series_info[13], end="  ")
    print(series_info[14], end="  ")
    print(series_info[15])


if __name__ == '__main__':
    old_workbook = openpyxl.load_workbook("bilibili_bangumi_1.xlsx")
    all_workbook = openpyxl.load_workbook("bilibili_bangumi_all.xlsx")
    series_workbook = openpyxl.load_workbook("series.xlsx")
    old_sheet = old_workbook[old_workbook.sheetnames[0]]
    episode_sheet = all_workbook[all_workbook.sheetnames[0]]
    series_sheet = series_workbook[series_workbook.sheetnames[0]]

    # all_sheet.cell(1,1).alignment = Alignment(horizontal='center', vertical='center')
    # all_sheet.cell(1,2).alignment = Alignment(horizontal='center', vertical='center')
    # all_sheet.cell(1,3).alignment = Alignment(horizontal='center', vertical='center')
    # all_sheet.cell(1,4).alignment = Alignment(horizontal='center', vertical='center')
    # all_sheet.cell(1,5).alignment = Alignment(horizontal='center', vertical='center')
    # all_sheet.cell(1,6).alignment = Alignment(horizontal='center', vertical='center')
    # all_sheet.cell(1,7).alignment = Alignment(horizontal='center', vertical='center')
    # all_sheet.cell(1,8).alignment = Alignment(horizontal='center', vertical='center')
    # all_sheet.cell(1,9).alignment = Alignment(horizontal='center', vertical='center')
    # all_sheet.cell(1,10).alignment = Alignment(horizontal='center', vertical='center')
    # all_sheet.cell(1,11).alignment = Alignment(horizontal='center', vertical='center')
    # all_sheet.cell(1,12).alignment = Alignment(horizontal='center', vertical='center')
    # all_sheet.cell(1,13).alignment = Alignment(horizontal='center', vertical='center')
    # all_sheet.cell(1,14).alignment = Alignment(horizontal='center', vertical='center')
    # all_sheet.cell(1,15).alignment = Alignment(horizontal='center', vertical='center')
    # all_sheet.cell(1,16).alignment = Alignment(horizontal='center', vertical='center')
    # all_sheet.cell(1,17).alignment = Alignment(horizontal='center', vertical='center')
    # all_sheet.cell(1,18).alignment = Alignment(horizontal='center', vertical='center')
    row_count = old_sheet.max_row
    row = episode_sheet.max_row + 1
    try:
        start_id = int(episode_sheet.cell(episode_sheet.max_row, 1).value) + 1
    except:
        start_id = 0

    try:
        i = 2
        while i < row_count + 1:
            ep_id = int(old_sheet.cell(i, 1).value)  # 1
            if ep_id < start_id:
                i += 1
                continue

            content = get_content(ep_id)
            if content == "412":
                all_workbook.save("bilibili_bangumi_all.xlsx")
                series_workbook.save("series.xlsx")
                while content == "412":
                    print("error")
                    time.sleep(300)
                    content = get_content(ep_id)
            if "出错啦! - bilibili.com" in content:
                i += 1
                print("missing")
                continue
            # info = eval(content.split("<script>window.__INITIAL_STATE__=")[1].split(";(function(){")[0])
            try:
                info = content.split("<script>window.__INITIAL_STATE__=")[1].split(";(function(){")[0]
            except IndexError:
                continue
            info = info.replace("false", "False")
            info = info.replace("null", "None")
            info = info.replace("true", "True")
            info = eval(info)

            ss = {}
            for series in info["ssList"]:
                if series["id"] == info["mediaInfo"]["ssId"]:
                    ss = series
                    break
            media_stat_coins = info["mediaInfo"]["stat"]["coins"]  # 13
            media_stat_danmakus = info["mediaInfo"]["stat"]["danmakus"]  # 12
            media_stat_favorites = info["mediaInfo"]["stat"]["favorites"]  # 14
            media_stat_reply = info["mediaInfo"]["stat"]["reply"]  # 15
            media_stat_share = info["mediaInfo"]["stat"]["share"]  # 16
            media_title = info["mediaInfo"]["title"]
            media_series = info["mediaInfo"]["series"]  # 6
            media_ssType = info["mediaInfo"]["ssTypeFormat"]["name"]
            media_pub_time = info["mediaInfo"]["pub"]["time"]  # 8
            media_rating_score = info["mediaInfo"]["rating"]["score"]  # 17
            media_rating_count = info["mediaInfo"]["rating"]["count"]  # 18
            media_newestEp_desc = info["mediaInfo"]["newestEp"]["desc"]  # 9
            # ep_title = info["epInfo"]["title"]
            ep_titleFormat = info["epInfo"]["titleFormat"]  # 3
            ep_longTitle = info["epInfo"]["longTitle"]  # 4
            try:
                ss_id = ss["id"]
                ss_title = ss["title"]
                ss_type = ss["pgcType"]
                ss_views = ss["views"]
                ss_follows = ss["follows"]
                ss_error = False
            except KeyError:
                ss_id = info["mediaInfo"]["ssId"]
                ss_title = media_title
                ss_type = info["mediaInfo"]["pgcType"]
                ss_views = info["mediaInfo"]["stat"]["views"]
                ss_follows = 0
                ss_error = True
            if ss_title == "":
                ss_title = media_title

            episode_sheet.cell(row, 1).value = ep_id
            episode_sheet.cell(row, 2).value = media_title
            episode_sheet.cell(row, 3).value = ep_titleFormat
            episode_sheet.cell(row, 4).value = ep_longTitle
            episode_sheet.cell(row, 5).value = ss_id
            episode_sheet.cell(row, 6).value = media_series
            # all_sheet.cell(row, 7).value = ss_type
            # all_sheet.cell(row, 8).value = media_pub_time
            # all_sheet.cell(row, 9).value = media_newestEp_desc
            # all_sheet.cell(row, 10).value = ss_views
            # all_sheet.cell(row, 11).value = ss_follows
            # all_sheet.cell(row, 12).value = media_stat_danmakus
            # all_sheet.cell(row, 13).value = media_stat_coins
            # all_sheet.cell(row, 14).value = media_stat_favorites
            # all_sheet.cell(row, 15).value = media_stat_reply
            # all_sheet.cell(row, 16).value = media_stat_share
            # all_sheet.cell(row, 17).value = media_rating_score
            # all_sheet.cell(row, 18).value = media_rating_count
            episode_sheet.cell(row, 1).alignment = Alignment(horizontal='center', vertical='center')
            episode_sheet.cell(row, 2).alignment = Alignment(horizontal='center', vertical='center')
            episode_sheet.cell(row, 3).alignment = Alignment(horizontal='center', vertical='center')
            episode_sheet.cell(row, 4).alignment = Alignment(horizontal='center', vertical='center')
            episode_sheet.cell(row, 5).alignment = Alignment(horizontal='center', vertical='center')
            episode_sheet.cell(row, 6).alignment = Alignment(horizontal='center', vertical='center')
            # all_sheet.cell(row, 7).alignment = Alignment(horizontal='center', vertical='center')
            # all_sheet.cell(row, 8).alignment = Alignment(horizontal='center', vertical='center')
            # all_sheet.cell(row, 9).alignment = Alignment(horizontal='center', vertical='center')
            # all_sheet.cell(row, 10).alignment = Alignment(horizontal='center', vertical='center')
            # all_sheet.cell(row, 11).alignment = Alignment(horizontal='center', vertical='center')
            # all_sheet.cell(row, 12).alignment = Alignment(horizontal='center', vertical='center')
            # all_sheet.cell(row, 13).alignment = Alignment(horizontal='center', vertical='center')
            # all_sheet.cell(row, 14).alignment = Alignment(horizontal='center', vertical='center')
            # all_sheet.cell(row, 15).alignment = Alignment(horizontal='center', vertical='center')
            # all_sheet.cell(row, 16).alignment = Alignment(horizontal='center', vertical='center')
            # all_sheet.cell(row, 17).alignment = Alignment(horizontal='center', vertical='center')
            # all_sheet.cell(row, 18).alignment = Alignment(horizontal='center', vertical='center')

            ss_info = [
                ss_id,
                media_series,
                ss_title,
                ss_type,
                media_pub_time,
                media_newestEp_desc,
                ss_views,
                ss_follows,
                media_stat_danmakus,
                media_stat_coins,
                media_stat_favorites,
                media_stat_reply,
                media_stat_share,
                media_rating_score,
                media_rating_count,
                media_ssType
            ]
            save_series(series_sheet, ss_info)

            row += 1

            if ss_error: print("series_error", end="  ")
            print(ep_id, end="  ")
            print(media_title, end="  ")
            print(ep_titleFormat, end="  ")
            print(ep_longTitle, end="  ")
            print(ss_id, end="  ")
            print(media_series)

            i += 1
    except KeyboardInterrupt:
        print("KeyboardInterrupt")
    finally:
        old_workbook.close()
        all_workbook.save("bilibili_bangumi_all.xlsx")
        all_workbook.close()
        series_workbook.save("series.xlsx")
        series_workbook.close()
