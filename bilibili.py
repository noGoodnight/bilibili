import requests
import openpyxl
import time
from openpyxl.styles import Alignment


def get_title(ep_id: int) -> str:
    data = requests.get("https://www.bilibili.com/bangumi/play/ep" + str(ep_id))
    content = data.content.decode("utf8")
    if "由于触发哔哩哔哩安全风控策略，该次访问请求被拒绝。" in content:
        return "412"
    else:
        return content.split("<title>")[1].split("</title")[0]


if __name__ == '__main__':
    # try:
    #     out_file = openpyxl.load_workbook("bilibili_bangumi.xlsx")
    #     sheet = out_file.get_sheet_by_name(out_file.sheetnames[0])
    # except:
    out_file = openpyxl.load_workbook("bilibili_bangumi_2.xlsx")
    sheet_id = out_file.sheetnames[0]
    sheet = out_file[sheet_id]
    sheet.column_dimensions['B'].width = 50
    sheet_id = int(sheet_id)
    next_sheet_id = sheet_id + 100000

    # row_count = sheet.max_row
    # # if row_count == 1:
    # sheet.cell(1, 1).value = "id"
    # sheet.cell(1, 1).alignment = Alignment(horizontal='center', vertical='center')
    # sheet.cell(1, 2).value = "name"
    # sheet.cell(1, 2).alignment = Alignment(horizontal='center', vertical='center')
    # sheet.cell(1, 3).value = "episode"
    # sheet.cell(1, 3).alignment = Alignment(horizontal='center', vertical='center')
    row_count = sheet.max_row + 1

    try:
        for i in range(sheet_id, min(sheet_id + 100000, 200000)):
            next_sheet_id = i
            # time.sleep(0.25)
            print(i, end="        ")
            title = get_title(i)
            if title == "412":

                print("ERROR", end="        ")
                sheet.title = str(next_sheet_id)
                out_file.save("bilibili_bangumi_2.xlsx")
                title = get_title(i)

                while title == "412":
                    time.sleep(600)
                    title = get_title(i)
                    print("ERROR", end="        ")

                # break
            if title == "出错啦! - bilibili.com":
                print()
                continue
            if "：" in title:
                title_info = title.split("：")
                title_info[0] = "：".join(title_info[0:len(title_info) - 1])
                title_info[1] = title_info[len(title_info) - 1]
                bangumi_name = title_info[0]
                bangumi_episode = title_info[1].split("_")[0]
            else:
                print("special", end="        ")
                title_info = title.split("_")
                title_info[1] = ""
                bangumi_name = title_info[0]
                bangumi_episode = title_info[1].split("_")[0]
            sheet.cell(row_count, 1).value = i
            sheet.cell(row_count, 1).alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row_count, 2).value = bangumi_name
            sheet.cell(row_count, 2).alignment = Alignment(horizontal='center', vertical='center')
            sheet.cell(row_count, 3).value = bangumi_episode
            sheet.cell(row_count, 3).alignment = Alignment(horizontal='center', vertical='center')
            row_count += 1
            print(bangumi_name, end="        ")
            print(bangumi_episode)
    except KeyboardInterrupt:
        print("KeyboardInterrupt")
    except IndexError:
        print(title_info)
    finally:
        sheet.title = str(next_sheet_id)
        out_file.save("bilibili_bangumi_2.xlsx")
