import openpyxl


def diff(sheet, gap) -> int:
    global sheetall
    rowcount = sheet.max_row
    for i in range(2, rowcount):
        value = sheet.cell(i, 1).value
        valueall = sheetall.cell(i + gap, 1).value
        if value != valueall:
            print(sheet.cell(i, 1).value)
            print(sheet.cell(i, 2).value)
            gap -= 1
    gap = rowcount + gap - 1
    return gap


if __name__ == '__main__':
    wb1 = openpyxl.load_workbook("bilibili_bangumi_1.xlsx")
    wb0 = openpyxl.load_workbook("bilibili_bangumi_0.xlsx")
    wball = openpyxl.load_workbook("bilibili_bangumi_all.xlsx")
    sheet1 = wb1[wb1.sheetnames[0]]
    sheet0 = wb0[wb0.sheetnames[0]]
    sheetall = wball[wball.sheetnames[0]]
    rowcount1 = sheet1.max_row
    rowcount0 = sheet0.max_row
    rowcountall = sheetall.max_row
    gap = 0
    gap = diff(sheet0, gap)
    gap = diff(sheet1, gap)
